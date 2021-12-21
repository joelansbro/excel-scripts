# iterate dimensions columns and if the weights are decimalised rounds them up
# sorts them in following order =MAX =MIN =MEDIAN (length height width)
import openpyxl
from openpyxl import load_workbook

# gets col header based on parameter
def findCol(col_name, ws):
    for row in range(1, ws.max_column):
        if(ws.cell(1,row).value == col_name):
            return ws.cell(1,row)

# sorts the length height and width into max, min, median
def sortDims(length, height, width):
    if(length < width):
        length, width = width, length
    if(length < height):
        length, height = height, length
    if(width < height):
        width, height = height, width
    return[length, height, width]

# this dim sort is being fed coordinates eg ['C1'], ['D1'], ['E1']
# take the column and iterate down, sorting the variables in the order defined in sortDims
def dimSort(length, height, width, ws):
    hand = []
    print(length, height, width)
    for row in range(2,ws.max_row):
        
        lengthCell = str(length) + str(row)
        heightCell = str(height) + str(row)
        widthCell = str(width) + str(row)
        print(lengthCell)

        hand.append(ws[lengthCell].value)
        hand.append(ws[heightCell].value)
        hand.append(ws[widthCell].value)

        sortedRow = sortDims(hand[0], hand[1], hand[2])
        ws[lengthCell] = sortedRow[0]
        ws[heightCell] = sortedRow[1]
        ws[widthCell] = sortedRow[2]
        print("Sorted row " + str(row) +"Variables are now: "+ str(hand))
        hand = [] 

if __name__ in '__main__':
    wb = openpyxl.load_workbook(input("Input the absolute path of file: \n") + ".xlsx", read_only=False)
    ws = wb.active

    print("This programme will sort the Length, Height, and Width columns where\n")
    print("Lengh is the largest value, Height is the smallest, and Width is median.")

    VAR_length = findCol(input("Type the name of the Length column \n"), ws)
    VAR_height = findCol(input("Type the name of the Height column \n"), ws)
    VAR_width = findCol(input("Type the name of the Width column \n"), ws)

    print("Length = " + VAR_length.value)
    print("Height = " + VAR_height.value)
    print("Width = " + VAR_width.value)
    dimSort(VAR_length.column_letter, VAR_height.column_letter, VAR_width.column_letter, ws)
    
    wb.save(filename = input("Sorting of columns complete, save down your " + ".xlsx"))