# WIP

from os import read
import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

# file = input("Enter file name here: \n") + ".xlsx"
# wb = openpyxl.load_workbook(file, read_only=False)
# ws = wb.active

 
# stores the range products and dumps when complete the duplicate check
rangeDict = {
    "SKU1": '', # the first product of the range
    "SKU2": '' # the second product of the range
}

sheetRange =[]

rangeSKU = ['A2']
rangeSKUTotal = 'A'+str(ws.max_row)

column_entry = input("What is the start of the options columns?")


# to get the columns to iterate through
def getStartColumn(column_entry):
    for col in range(1, ws.max_column+1):
        if (ws.cell == column_entry):
            for cell in col:
                sheetRange.append(cell.value)

# find column based on name
def findCol(col_name, ws):
    for row in range(1, ws.max_column):
        if(ws.cell(1,row).value == col_name):
            return ws.cell(1,row).value


# to iterate through the rows
def iterateCols(column selection){

}

must store the values of the products within the range inside another data structure to compare them
def storethevaluesoftherangeinadataset(){
    # append the range products to the matrixdict, concatenating the cell values together
}


def compare the values(listMatrix){
# check if values of lists are equal
# need a master list to contain lists of all the products in a range, then once ran the check function
# find the duplicate list SKU and highlight the SKU,
# then empty the listmatrix and go again

    for list in listMatrix:
          
}


# store the range that we are searching for duplicates within
def getRange(){
    sheetRange = 
}


srcColEnd = 'G'+str(ws.max_row)
queryCol = 'G2:'+srcColEnd

red_fill = PatternFill(start_color="F20000", end_color='F20000', fill_type='solid')
ws.conditional_formatting.add(queryCol, FormulaRule(formula=['COUNTIF(G:G,G2)>1'], stopIfTrue=True, fill=red_fill))


newSave = input("Save new book as: ")
wb.save(newSave + '.xlsx')


if __name__ in '__main__':
    wb = openpyxl.load_workbook(input("absolute path of file") + ".xlsx", read_only=False)
    ws = wb.active