# searches through different excel sheets within a folder and returns a list of the sheets with the input

import os
import fnmatch
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

def search(filename, keyword):
    wb = load_workbook(filename)

    for s in wb.get_sheet_names():
        sh = wb[s]

        for col in range(1, sh.max_column+1):
            for row_index in range(1, sh.max_row+1):
                if sh.cell(row=row_index, column=col).value == keyword:
                    print(filename)


def get_all_target(path_target):
    file_list = []

    for dir, dn, f in os.walk(path_target):
        
        for ext in ['xls', 'xlsx']:
            temp_list = fnmatch.filter(os.listdir(dir), '*.' + ext)
            
            for i in range(len(temp_list)):
                if len(temp_list[i]) > 0:
                    temp_list[i] = dir + '\\' + temp_list[i]
                    file_list.append(temp_list[i])

        return(file_list)

if __name__ in '__main__':
    keyword = input("Enter the value you are searching for")
    path_target = input("Enter the absolute path for the folder you are seacrhing")
    print("********************\n")

    files = get_all_target(path_target)
    for file in files:
        search(file, keyword)